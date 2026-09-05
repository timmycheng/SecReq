# -*- coding: utf-8 -*-
"""API 接口批量导入解析(#92): 粘贴文本 / xlsx / csv → 行数组(预览用, 不落库)。

两段式: 本模块只做「解析+逐行校验」, 确认导入由前端合并进清单后走既有
POST /api/projects/{id}/api-endpoints 整体保存 —— 解析与持久化解耦。
列约定(与前端模板说明一致): 名称, 方法, 路径, 需要认证, 公网暴露;
需要认证/公网暴露容错 是/否/true/false/1/0/y/n; 缺省分别为 是/否。
"""
import csv
import io
import json

import shared.constants as C

_BOOL_TRUE = {"是", "y", "yes", "true", "1"}
_BOOL_FALSE = {"否", "n", "no", "false", "0"}


def _parse_bool(raw: str, default: bool) -> tuple[bool | None, str | None]:
    text = (raw or "").strip().lower()
    if not text:
        return default, None
    if text in _BOOL_TRUE:
        return True, None
    if text in _BOOL_FALSE:
        return False, None
    return None, f"布尔值无法识别: {raw!r}(可用 是/否/true/false/1/0)"


def _parse_line(line: str, index: int) -> dict:
    """单行 → 行对象; 错误写入 error 字段(不抛出, 非法行不阻塞合法行)。"""
    row = {"index": index, "name": "", "method": "GET", "path": "",
           "auth_required": True, "public_exposed": False, "error": None}
    cells = [c.strip() for c in line.replace("\t", ",").split(",")]
    errors: list[str] = []
    if len(cells) < 3:
        row["error"] = f"列数不足({len(cells)}), 需要 名称,方法,路径[,需要认证,公网暴露]"
        return row
    row["name"], method, path = cells[0], cells[1], cells[2]
    if not row["name"]:
        errors.append("名称为空")
    if not path:
        errors.append("路径为空")
    else:
        row["path"] = path
    method = method.upper()
    if method not in C.HTTP_METHODS:
        errors.append(f"HTTP 方法非法: {cells[1]!r}(可用 {'/'.join(C.HTTP_METHODS)})")
    else:
        row["method"] = method
    if len(cells) >= 4:
        ok, err = _parse_bool(cells[3], True)
        if ok is None:
            errors.append(err)
        else:
            row["auth_required"] = ok
    if len(cells) >= 5:
        ok, err = _parse_bool(cells[4], False)
        if ok is None:
            errors.append(err)
        else:
            row["public_exposed"] = ok
    if errors:
        row["error"] = "; ".join(errors)
    return row


def parse_text(text: str) -> list[dict]:
    """粘贴文本: 每行一条, 逗号/Tab 分列; 空行与 # 注释行跳过。"""
    rows = []
    index = 0
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        index += 1
        rows.append(_parse_line(line, index))
    return rows


def _looks_like_header(cells: list[str]) -> bool:
    joined = "".join(str(c) for c in cells)
    return ("名称" in joined or "name" in joined.lower()) and ("方法" in joined or "method" in joined.lower())


def parse_xlsx(content: bytes) -> list[dict]:
    """xlsx 首个工作表: 每行一条(自动跳过表头行)。"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    index = 0
    header_skipped = False
    for excel_row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c) for c in excel_row]
        if not any(c.strip() for c in cells):
            continue
        if not header_skipped:
            header_skipped = True
            if _looks_like_header(cells):
                continue
        index += 1
        rows.append(_parse_line(",".join(cells), index))
    wb.close()
    return rows


def parse_csv(content: bytes) -> list[dict]:
    """csv/txt: 按文本行解析(与粘贴文本同一套规则)。"""
    text = content.decode("utf-8-sig", errors="replace")
    # csv 场景存在引号包裹字段时交给 csv.reader, 其余退化按行
    sample = text[:1024]
    if '"' in sample:
        rows: list[dict] = []
        index = 0
        header_skipped = False
        reader = csv.reader(io.StringIO(text))
        for cells in reader:
            line = ",".join(cells)
            if not line.strip() or line.strip().startswith("#"):
                continue
            if not header_skipped:
                header_skipped = True
                if _looks_like_header(cells):
                    continue
            index += 1
            rows.append(_parse_line(line, index))
        return rows
    return parse_text(text)


def parse_openapi(text: str) -> list[dict]:
    """OpenAPI 3 / Swagger 2 规范文件解析(#227), JSON 与 YAML 均可。

    接口行取 paths 的 名称(summary/operationId)/方法/路径; 需要认证按
    security 定义推断默认值 —— 操作级 security 覆盖全局, 全局 security
    为空列表视为无需认证, 其余(含未声明)默认需要认证。
    """
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        try:
            import yaml
        except ImportError as exc:  # pragma: no cover
            raise ValueError("YAML 解析依赖 pyyaml 未安装") from exc
        try:
            data = yaml.safe_load(text)
        except yaml.YAMLError as exc:
            raise ValueError(f"YAML 解析失败: {exc}") from exc
    if not isinstance(data, dict) or ("paths" not in data):
        raise ValueError("文件不是 OpenAPI/Swagger 规范(缺少 paths 定义)")

    def security_requires_auth(op: dict) -> bool:
        sec = op.get("security", data.get("security"))
        if sec is None:
            return True
        return len(sec) > 0  # 显式空列表 = 匿名可访问

    rows: list[dict] = []
    for path, methods in (data.get("paths") or {}).items():
        if not isinstance(methods, dict):
            continue
        for method, op in methods.items():
            up = str(method).upper()
            if up not in C.HTTP_METHODS:
                continue  # parameters/get 之外的键(如 parameters/servers)跳过
            if not isinstance(op, dict):
                op = {}
            name = str(op.get("summary") or op.get("operationId") or f"{up} {path}")
            rows.append({
                "index": len(rows) + 1,
                "name": name[:200],
                "method": up,
                "path": str(path),
                "auth_required": security_requires_auth(op),
                "public_exposed": False,
                "error": None,
            })
    return rows


def parse_upload(filename: str, content: bytes) -> list[dict]:
    """按扩展名分派: xlsx → parse_xlsx; json/yaml → OpenAPI/Swagger(#227); csv/txt → parse_csv。"""
    lowered = (filename or "").lower()
    if lowered.endswith(".xlsx") or lowered.endswith(".xls"):
        return parse_xlsx(content)
    if lowered.endswith((".json", ".yaml", ".yml")):
        text = content.decode("utf-8-sig", errors="replace")
        try:
            return parse_openapi(text)
        except ValueError as exc:
            # 非规范文件给出可读报错行(不抛出, 与逐行错误同形态)
            return [{"index": 1, "name": "", "method": "GET", "path": "",
                     "auth_required": True, "public_exposed": False,
                     "error": str(exc)}]
    return parse_csv(content)
