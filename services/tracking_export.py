# -*- coding: utf-8 -*-
"""Excel 需求跟踪表导出(DESIGN.md 模块5)。

字段口径: req_id、需求描述、优先级、责任方、建议阶段、验收标准、状态(默认open)、备注。
备注列携带追溯信息(来源中文标签 ← trigger_reason), 满足"需求必须可追溯到输入"约束。
第二个工作表给出 Jira 外部系统导入(CSV)的字段映射建议。
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import shared.constants as C

# (列表头, 列宽, 对齐)
_COLUMNS = [
    ("req_id", 16, "left"),
    ("需求描述", 52, "left"),
    ("优先级", 8, "center"),
    ("责任方", 10, "center"),
    ("建议阶段", 12, "center"),
    ("验收标准", 52, "left"),
    ("合规依据", 36, "left"),
    ("状态", 10, "center"),
    ("备注", 44, "left"),
]

_JIRA_HINTS = [
    ["SecReq 安全需求跟踪表 — Jira 导入字段映射建议"],
    [""],
    ["1. 本表可另存为 CSV 后使用 Jira「外部系统导入(External System Import)」导入;"],
    ["2. 建议映射:"],
    ["      需求描述 → Summary(问题标题);"],
    ["      验收标准 + 需求全文 → Description(描述);"],
    ["      优先级 → Priority(Priority Name 需在 Jira 中预置同名值);"],
    ["      建议阶段 → Labels(标签, 如 设计阶段/开发阶段/测试阶段);"],
    ["      状态 → 建议不直接映射 Status, 导入后按团队流程流转;"],
    ["3. 备注=需求追溯信息: 来源(中文) ← 触发原因。"],
]


#: 漏洞清单工作表(v2.2.0): 合规通报常要求 CNNVD 编号, 与需求跟踪表一并交付
_VULN_COLUMNS = [
    ("等级", 10, "center"),
    ("CVE", 22, "left"),
    ("CNNVD", 22, "left"),
    ("中文等级", 10, "center"),
    ("组件", 26, "left"),
    ("受影响范围", 22, "left"),
    ("修复版本", 14, "left"),
    ("数据来源", 14, "center"),
    ("简述", 44, "left"),
]

_VULN_SOURCE_LABELS = {
    "osv_local": "本地漏洞库",
    "osv_online": "OSV.dev 在线",
    "sca": "行内 SCA",
}


def build_tracking_workbook(requirements: list, vulnerabilities: list | None = None) -> Workbook:
    """requirements 为 SecurityRequirement 记录列表, 返回内存工作簿。

    vulnerabilities 可选: 传入时追加「漏洞清单」工作表(含 CNNVD 编号与数据来源)。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "跟踪表"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")

    for col, (title, width, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    priority_order = {p: i for i, p in enumerate(["critical", "high", "medium", "low"])}
    rows = sorted(
        requirements,
        key=lambda r: (priority_order.get(r.priority, 9), r.category, r.req_id),
    )
    for idx, req in enumerate(rows, start=2):
        trace = f"{getattr(req, 'source_label', None) or req.source_entity_type} ← {req.trigger_reason}"
        status_label = C.label(C.REQUIREMENT_STATUS, req.status or "open", "待处理")
        basis = "; ".join(
            f"《{ref.get('file', '')}》{ref.get('clause', '')}"
            if ref.get("clause") else f"《{ref.get('file', '')}》"
            for ref in (getattr(req, "regulatory_ref", None) or [])
            if ref.get("file")
        )
        values = [
            req.req_id,
            f"【{C.label(C.REQUIREMENT_PRIORITY_LABELS, req.priority)}】{req.title}\n{req.description}",
            C.label(C.REQUIREMENT_PRIORITY_LABELS, req.priority),
            getattr(req, "owner", None) or "安全/开发团队",
            C.label(C.REQUIREMENT_PHASES, req.suggested_phase, req.suggested_phase),
            req.acceptance_criteria,
            basis or "—",
            status_label,
            trace,
        ]
        for col, value in enumerate(values, start=1):
            _, _, align = _COLUMNS[col - 1]
            cell = ws.cell(row=idx, column=col, value=value)
            cell.alignment = Alignment(
                horizontal=align, vertical="top", wrap_text=(align == "left"))
            if isinstance(value, str) and "\n" in value:
                cell.alignment = Alignment(
                    horizontal=align, vertical="top", wrap_text=True)

    if vulnerabilities is not None:
        _append_vuln_sheet(wb, vulnerabilities)

    hints = wb.create_sheet("Jira导入说明")
    hints.column_dimensions["A"].width = 100
    for row_idx, hint_row in enumerate(_JIRA_HINTS, start=1):
        cell = hints.cell(row=row_idx, column=1, value=hint_row[0] if len(hint_row) == 1 else "")
        cell.alignment = Alignment(vertical="top", wrap_text=False)
    return wb


def _append_vuln_sheet(wb: Workbook, vulnerabilities: list) -> None:
    """追加漏洞清单工作表; 表头样式与跟踪表保持一致。"""
    ws = wb.create_sheet("漏洞清单")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for col, (title, width, _) in enumerate(_VULN_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    severity_order = {s: i for i, s in enumerate(["critical", "high", "medium", "low"])}
    rows = sorted(
        vulnerabilities,
        key=lambda v: (severity_order.get(v.severity, 9), getattr(v, "component_name", "") or "", v.cve_id),
    )
    for idx, v in enumerate(rows, start=2):
        version = getattr(v, "component_version", "") or ""
        comp = f"{getattr(v, 'component_name', '')}@{version}".strip("@")
        values = [
            C.label(C.SEVERITY_LABELS, v.severity),
            v.cve_id,
            getattr(v, "cnnvd_id", None) or "—",
            getattr(v, "cn_severity", None) or "—",
            comp or "—",
            v.affected_range or "—",
            v.fix_version or "未发布",
            _VULN_SOURCE_LABELS.get(getattr(v, "source", "") or "", "本地漏洞库"),
            v.summary or "",
        ]
        for col, value in enumerate(values, start=1):
            _, _, align = _VULN_COLUMNS[col - 1]
            ws.cell(row=idx, column=col, value=value).alignment = Alignment(
                horizontal=align, vertical="top", wrap_text=(align == "left"))


def tracking_xlsx_bytes(requirements: list, vulnerabilities: list | None = None) -> bytes:
    """工作簿序列化为字节流(API 直接返回)。"""
    buffer = BytesIO()
    build_tracking_workbook(requirements, vulnerabilities).save(buffer)
    return buffer.getvalue()
