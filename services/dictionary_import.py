# -*- coding: utf-8 -*-
"""数据字典导入与自动分级(走查整改 Step-4)。

开发把数据字典(表/字段清单)粘贴或上传上来, 本模块解析出行结构并按字段名
模式库自动推断:
- 字段级: JR/T 0197 分级建议、PII/敏感PII 标记、是否需要脱敏(含脱敏规则);
- 表级/资产级: 取所含字段的最高分级。

仅返回建议(不落库); 用户在前端确认/修正后走现有 saveDataAssets 保存。
"""
import io
import re

import shared.constants as C

# 字段级敏感模式: (正则, 建议分级, is_pii, is_sensitive_pii)
# 顺序即优先级: 先命中先得(敏感 PII > 一般 PII > 财务业务 > 默认)。
_FIELD_PATTERNS: list[tuple[str, str, str]] = [
    # (pattern, kind, level)
    (r"(支付密码|交易密码|登录口令|password|passwd|pwd)", "password", "4级_C3鉴别信息"),
    (r"(cvn2?|cvv|磁道|card_pin|pin_?block)", "card_pin", "4级_C3鉴别信息"),
    (r"(指纹|人脸|虹膜|声纹|生物特征|biometric)", "biometric", "4级_C3鉴别信息"),
    (r"(证书私钥|private_?key|secret_?key)", "private_key", "4级_C3鉴别信息"),
    (r"(身份证|证件号|identity|id_?card|护照号|军官证)", "id_card", "4级_C3鉴别信息"),
    (r"(银行卡|卡号|bank_?card|card_?num|账号|account_?no|acct)", "bank_card", "3级_C2主要信息"),
    (r"(手机|电话|mobile|phone)", "phone_number", "3级_C2主要信息"),
    (r"(邮箱|email)", "email", "3级_C2主要信息"),
    (r"(姓名|customer_?name|user_?name|户名)", "name", "3级_C2主要信息"),
    (r"(住址|地址|address)", "address", "3级_C2主要信息"),
    (r"(生日|出生日期|birth)", "birth", "3级_C2主要信息"),
    (r"(余额|交易|流水|金额|amount|balance|账单|账务)", "finance", "3级_C2主要信息"),
    (r"(证件类型|证件有效期|国籍|民族|性别|婚姻)", "kyc", "3级_C2主要信息"),
]

_MASK_FIELD_PATTERNS = dict(C.MASK_FIELD_PATTERNS)
_MASK_FIELD_PATTERNS["password"] = "密码明文禁止留存, 传输须加密"
_MASK_FIELD_PATTERNS["card_pin"] = "CVN/PIN 禁止存储与展示"
_MASK_FIELD_PATTERNS["biometric"] = "生物特征模板禁止明文导出"
_MASK_FIELD_PATTERNS["private_key"] = "私钥禁止明文落盘"
_MASK_FIELD_PATTERNS["finance"] = "金额/流水按业务需要脱敏展示"
_MASK_FIELD_PATTERNS["kyc"] = "证件要素部分遮蔽"

_DEFAULT_LEVEL = "2级_C1次要信息"
_DEFAULT_TYPE = "VARCHAR"
_SPLIT_RE = re.compile(r"[\t|,，;；]+")
_MARKDOWN_SEP_RE = re.compile(r"^\|?[\s:|-]+\|?\s*$")
_HEADER_HINTS = ("表名", "字段", "table", "column", "field", "列名")


def parse_dictionary_text(text: str) -> list[dict]:
    """解析粘贴的表/字段清单文本。支持 TSV、竖线、逗号分隔或 markdown 表格。

    每行 2 列以上: [表名, 字段名, 类型?]; 1 列: 字段名(或 表.字段 点分)。
    """
    rows: list[dict] = []
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or _MARKDOWN_SEP_RE.match(line):
            continue
        cells = [c.strip() for c in _SPLIT_RE.split(line) if c.strip()]
        if len(cells) == 1 and "." in cells[0] and not cells[0].startswith("#"):
            table, _, field = cells[0].partition(".")
            rows.append({"table": table.strip(), "field": field.strip(), "type": ""})
            continue
        if len(cells) >= 2:
            if any(h in cells[0].lower() for h in _HEADER_HINTS):
                continue  # 表头行
            rows.append({
                "table": cells[0],
                "field": cells[1],
                "type": cells[2] if len(cells) >= 3 else "",
            })
    return rows


def parse_dictionary_xlsx(content: bytes) -> list[dict]:
    """解析上传的 xlsx(取第一个 Sheet, 前 3 列按 表/字段/类型)。"""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(max_col=3, values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cells = [c for c in cells if c]
        if not cells:
            continue
        if i == 0 and any(h in " ".join(cells).lower() for h in _HEADER_HINTS):
            continue
        if len(cells) >= 2:
            rows.append({
                "table": cells[0],
                "field": cells[1],
                "type": cells[2] if len(cells) >= 3 else "",
            })
    return rows


def classify_field(field_name: str) -> dict:
    """字段名 → 分级/PII/脱敏建议。"""
    name = (field_name or "").strip()
    for pattern, kind, level in _FIELD_PATTERNS:
        if re.search(pattern, name, re.IGNORECASE):
            sensitive = level.startswith("4级")
            mask_rule = None
            if kind in _MASK_FIELD_PATTERNS:
                mask_rule = _MASK_FIELD_PATTERNS[kind]
            elif kind in _MASK_FIELD_PATTERNS:
                mask_rule = C.MASK_RULES.get(kind)
            return {
                "classification": level,
                "is_pii": True,
                "is_sensitive_pii": sensitive,
                "need_encrypt": sensitive,
                "need_mask": kind in _MASK_FIELD_PATTERNS,
                "mask_rule": mask_rule,
                "matched_kind": kind,
            }
    return {
        "classification": _DEFAULT_LEVEL,
        "is_pii": False,
        "is_sensitive_pii": False,
        "need_encrypt": False,
        "need_mask": False,
        "mask_rule": None,
        "matched_kind": None,
    }


_LEVEL_ORDER = list(C.DATA_LEVEL_ORDER.values())  # code → 5..1


def _level_rank(code: str) -> int:
    return C.level_rank(code)


def build_asset_suggestions(rows: list[dict]) -> list[dict]:
    """解析行 → 按表分组的资产建议(每张表一个资产, 分级取字段最高)。"""
    tables: dict[str, list[dict]] = {}
    order: list[str] = []
    for row in rows:
        table = row["table"] or "默认表"
        if table not in tables:
            tables[table] = []
            order.append(table)
        suggestion = classify_field(row["field"])
        tables[table].append({
            "field_name": row["field"],
            "field_type": row["type"] or _DEFAULT_TYPE,
            "need_encrypt": suggestion["need_encrypt"],
            "need_mask": suggestion["need_mask"],
            "mask_rule": suggestion["mask_rule"],
            "_level": suggestion["classification"],
            "_is_pii": suggestion["is_pii"],
            "_is_sensitive_pii": suggestion["is_sensitive_pii"],
            "_matched_kind": suggestion["matched_kind"],
        })

    assets: list[dict] = []
    for table in order:
        fields = tables[table]
        top = max(fields, key=lambda f: _level_rank(f["_level"]))
        has_biometric = any(f["_matched_kind"] == "biometric" for f in fields)
        assets.append({
            "name": table,
            "data_type": "biometric" if has_biometric else "business_data",
            "classification": top["_level"] or _DEFAULT_LEVEL,
            "c3_tag": bool(has_biometric and top["_level"].startswith("4级")),
            "is_pii": any(f["_is_pii"] for f in fields),
            "is_sensitive_pii": any(f["_is_sensitive_pii"] for f in fields),
            "storage_envs": ["db"],
            "cross_border_transfer": False,
            "tables": [{
                "table_name": table,
                "fields": [
                    {k: v for k, v in f.items() if not k.startswith("_")}
                    for f in fields
                ],
            }],
        })
    return assets
